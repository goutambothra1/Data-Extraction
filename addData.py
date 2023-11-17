import openpyxl
import os
from constants import LOGFILENAME,FILETYPE1,FILETYPE2,FILETYPE3,FILETYPE4
def addData(excel_file_path:str, file_name:str, file_type:str, levelofreview:str,  recent_programmer:str, recent_date:str,reviewer:str)->bool:

    # Load the Excel workbook
    logfilelocation = os.path.dirname(excel_file_path)
    workbook = openpyxl.load_workbook(excel_file_path)
    # Select the worksheet (assuming it's named "SDTM Programming Plan")
    sheet = workbook["Programming Plan"]        
    # Get the list of file names from column B
    file_names = [cell.value for cell in sheet["B"]]
    with open(os.path.join(logfilelocation,LOGFILENAME), 'a') as file:
        changes_detected=False
        # Check if the file_name already exists in the sheet
        if file_name in file_names:
            # If it exists, find its row number
            row_index = file_names.index(file_name) + 1
            new_record=file_type,file_name, levelofreview,  recent_programmer, recent_date,reviewer
            row = list(sheet.iter_rows( min_row=row_index,max_row=row_index,min_col=1, max_col=6, values_only=True))[0]
            if(new_record!=row):
            # Update the values in the existing row
                sheet.cell(row=row_index, column=1, value=file_type)
                sheet.cell(row=row_index, column=3, value=levelofreview)
                sheet.cell(row=row_index, column=4, value=recent_programmer)
                sheet.cell(row=row_index, column=5, value=recent_date)
                sheet.cell(row=row_index, column=6, value=reviewer)
                workbook.save(excel_file_path)
                file.write(f"Updated '{file_name}' in the Excel sheet. \n")
                file.close()
                changes_detected=True
                return changes_detected
        else:
            index=returnrowindex(excel_file_path)
            filetype1_index,filetype2_index,filetype3_index,filetype4_index=index
            # If it doesn't exist, add a new row with the provided values
            if file_type==FILETYPE1:
                next_row=filetype1_index+1
            elif file_type==FILETYPE2:
                next_row=filetype2_index+1
            elif file_type==FILETYPE3:
                next_row=filetype3_index+1
            elif file_type==FILETYPE4:
                next_row=filetype4_index+1
            sheet.insert_rows(next_row)
            # next_row = len(file_names) + 1
            sheet.cell(row=next_row, column=1, value=file_type)
            sheet.cell(row=next_row, column=2, value=file_name)
            sheet.cell(row=next_row, column=3, value=levelofreview)
            sheet.cell(row=next_row, column=4, value=recent_programmer)
            sheet.cell(row=next_row, column=5, value=recent_date)
            sheet.cell(row=next_row, column=6, value=reviewer)
            workbook.save(excel_file_path)
            file.write(f"Added '{file_name}' to the Excel sheet.\n")
            file.close()
            changes_detected=True
        return changes_detected
#to find the index of row which is last Type of that type
def returnrowindex(excel_file_path:str)->tuple[int,int,int,int]:
    filetype4_index=filetype3_index= filetype2_index=filetype1_index=1
    workbook=openpyxl.load_workbook(excel_file_path)
    sheet=workbook['Programming Plan']
    filetypes=[cell.value for cell in sheet["A"]]
    for index,filetype in enumerate(filetypes,1): 
        if filetype==FILETYPE1:
            filetype4_index=filetype3_index= filetype2_index=filetype1_index=index
        elif filetype==FILETYPE2:
            filetype4_index=filetype3_index=filetype2_index=index
        elif filetype==FILETYPE3:
            filetype4_index=filetype3_index=index
        elif filetype==FILETYPE4:
            filetype4_index=index
    return filetype1_index,filetype2_index,filetype3_index,filetype4_index

def removerows(excel_file_path:str,files_in_folder:list)->bool:
    changes_detected=False
    logfilelocation = os.path.dirname(excel_file_path)
    workbook=openpyxl.load_workbook(excel_file_path)
    sheet=workbook['Programming Plan']
    file_names_in_excel=[cell.value for cell in sheet["B"]][1:]
    for index,file_name in enumerate(file_names_in_excel,2):
        if file_name not in files_in_folder:
            with open (os.path.join(logfilelocation,LOGFILENAME),'a') as logfile:
                logfile.write(f"Deleted {file_name} from the Excel sheet . \n")
            logfile.close()
            sheet.delete_rows(index,1)
            workbook.save(excel_file_path)
            changes_detected = True 
    return changes_detected

        
            

    



    
    
    
    

